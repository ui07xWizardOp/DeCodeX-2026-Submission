import openpyxl

file_path = r'c:\Users\KIIT0001\Desktop\projects\DeCodeX\Response from agents\Response3attachment.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        
        # Get head
        # We handle potential huge sheets by limiting rows
        rows = list(ws.iter_rows(max_row=5, values_only=True))
        if rows:
            headers = rows[0]
            print(f"Columns: {list(headers)}")
            for i, row in enumerate(rows[1:], 1):
                if any(cell is not None for cell in row):
                    print(f"  Row {i}: {list(row)}")
except Exception as e:
    print(f"Error: {e}")
