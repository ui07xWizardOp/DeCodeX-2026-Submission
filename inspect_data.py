import openpyxl

file_path = r'c:\Users\KIIT0001\Desktop\projects\DeCodeX\DecodeX_VoltRide_Dataset.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Get rows
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    if rows:
        headers = rows[0]
        print(f"Columns ({len(headers)}): {list(headers)}")
        print("\nSample rows:")
        for i, row in enumerate(rows[1:], 1):
            if any(cell is not None for cell in row):
                print(f"  Row {i}: {list(row)}")
    
    # Count total rows properly
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            count += 1
    print(f"\nTotal non-empty rows: {count}")

wb.close()
