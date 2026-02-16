import openpyxl
import datetime

file_path = r'c:\Users\KIIT0001\Desktop\projects\DeCodeX\DecodeX_VoltRide_Dataset.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

with open(r'c:\Users\KIIT0001\Desktop\projects\DeCodeX\dataset_metadata.txt', 'w') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n=== Sheet: {sheet_name} ===\n")
        
        # Get rows
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        if rows:
            headers = rows[0]
            f.write(f"Columns: {list(headers)}\n")
            if len(rows) > 1:
                f.write(f"Sample Row: {list(rows[1])}\n")
        
        # Count rows
        count = 0
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                count += 1
        f.write(f"Total non-empty rows: {count}\n")

wb.close()
print("Metadata dumped to dataset_metadata.txt")
