import openpyxl

file_path = r'c:\Users\KIIT0001\Desktop\projects\DeCodeX\DecodeX_VoltRide_Dataset.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Get rows
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    if rows:
        headers = rows[0]
        print(f"Columns: {list(headers)}")
        if len(rows) > 1:
            print(f"Row 1: {list(rows[1])}")

wb.close()
