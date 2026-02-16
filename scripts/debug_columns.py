import openpyxl

DATA_FILE = "DecodeX_VoltRide_Dataset.xlsx"

print(f"Inspecting {DATA_FILE}...")
wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)
sheet = wb.active

print("--- Header (Row 1) ---")
for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    for idx, val in enumerate(row):
        print(f"{idx}: {val}")

print("\n--- Row 2 ---")
for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
    for idx, val in enumerate(row):
        print(f"{idx}: {val} (Type: {type(val)})")

print("\n--- Row 3 ---")
for row in sheet.iter_rows(min_row=3, max_row=3, values_only=True):
    for idx, val in enumerate(row):
        print(f"{idx}: {val} (Type: {type(val)})")
