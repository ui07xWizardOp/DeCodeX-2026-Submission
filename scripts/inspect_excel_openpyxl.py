import openpyxl
import os

print("Inspecting Excel with Openpyxl...")
data_file = "DecodeX_VoltRide_Dataset.xlsx"

try:
    wb = openpyxl.load_workbook(data_file, read_only=True, data_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    
    # Get Headers
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            headers.append(cell.value)
            
    print(f"Headers: {headers}")
    
    # Preview first row of data
    first_row = []
    for row in sheet.iter_rows(min_row=2, max_row=2):
        for cell in row:
            first_row.append(cell.value)
            
    print(f"First Row Data: {first_row}")
    
except Exception as e:
    print(f"Error: {e}")
